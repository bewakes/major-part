from openpyxl import *
import json
import sys
import os
import re
from subprocess import Popen, PIPE

def parse_vertical(sheet, metadata):
    data = {}
    data['columns'] = ['VDC/MUNICIPALITY']+metadata['ATTRIBUTES']
    data['rows'] = []
    maxrow = len(list(sheet.rows))
    cnt = 0
    for rownum in range(metadata['DATA_START'], maxrow):
        cnt+=1
        row = []
        vdcmun = sheet.cell(metadata['VDC_MUNICIPALITY']+str(rownum)).value
        if vdcmun==None:
            break
        row.append(vdcmun.upper())
        brk = True
        for column in metadata['COLUMNS']:
            cellvalue = sheet.cell(column+str(rownum)).value
            #print(cellvalue)
            if cellvalue is not None:
                brk = False
            row.append(cellvalue)
        data['rows'].append(row)
        if brk: break
    #print("rows evaluated: ", cnt)
    return data

def parse_horizontal(sheet, metadata):
    # merge vertical and horizontal attributes
    data = {}
    data['columns'] = ['VDC/MUNICIPALITY']
    data['rows'] = []
    for attr in metadata['ATTRIBUTES']:
        for vert in metadata['VERT_ATTRIBUTES']:
            data['columns'].append(attr+'_'+vert.upper().replace(' ','_'))

    maxrow = len(list(sheet.rows))
    #print('maxrow : ', maxrow)
    rownum = metadata['DATA_START']-1
    while rownum <= maxrow:
        row = []
        vdcmuncol = metadata['VDC_MUNICIPALITY']
        coord = chr(ord(vdcmuncol)-1)+str(rownum)
        temp = sheet.cell(coord).value
        #print(coord,temp)
        if temp==None or temp.strip()=='':
            rownum+=1
            continue
        else: # first row has vdc/mun name continue with others
            row.append(temp.upper())
            for column in metadata['COLUMNS']:
                for (i, attr) in enumerate(metadata['VERT_ATTRIBUTES']):
                    row.append(sheet.cell(column+str(rownum+i+1)).value)
            rownum+=len(metadata['VERT_ATTRIBUTES'])
            data['rows'].append(row)
    return data

parsedata = {'VERTICAL':parse_vertical, 'HORIZONTAL':parse_horizontal}

def parsefile(filename, jsonfile):
    workbook = load_workbook(filename=filename, read_only=True, data_only=True)
    sheetsinfo = json.load(open(jsonfile, 'r'))

    district_name = re.match(r'.*/(.+)\.xlsx', filename).group(1)
    out, error = Popen('mkdir districts_csv/'+district_name.replace(' ', '\ '),stdout=PIPE, stderr=PIPE, shell=True).communicate()
    print('error: ', error)
    if error != b'':
        print('CONTINUE TO NEXT')
        return
    # iterate through sheets
    for sheetinfo in sheetsinfo['SHEETS']:
        #print(sheetinfo['METADATA']['TITLE'])
        sheet = workbook.get_sheet_by_name(sheetinfo['NAME'])
        data = parsedata[sheetinfo['TYPE']](sheet, sheetinfo['METADATA'])
        write_to_csv(district_name+'/'+sheetinfo['METADATA']['TITLE'], data)
        #print(data['columns'])
        #for row in data['rows']:
            #print(row)
    #print('done')

def write_to_csv(filename, data):
    f = open('districts_csv/'+filename, 'w')
    f.write(','.join(data['columns'])+'\n')
    for row in data['rows']:
        f.write(','.join([str(x) for x in row])+'\n')
    f.close()


def main():
    try:
        folder = sys.argv[1]
        if folder[:-1]!='/':
            folder+='/'
        parsedatapath = sys.argv[2]
    except Exception:
        print("provide folder location and parse file location:")
        print("Example: ./parse.py <folder> <parsefilepath>")
    output = os.popen("ls "+folder+" -p | grep -v /").read()

    os.popen('mkdir districts_csv')
    filenames = [x for x in output.split('\n') if x!='']
    for name in filenames:
        print(name)
        parsefile(folder+name, parsedatapath)

if __name__=="__main__":
    main()
